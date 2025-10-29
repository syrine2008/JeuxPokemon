import termcolor
import openpyxl 
from models.Arene import Arene
import random
from tabulate import tabulate

#-------------------------------------
#   get nb des arenes 
#------------------------------------
def get_nb_arenes():
    nb_arenes=0
    #recuperer les arenes existantes 
    wb = openpyxl.load_workbook('arenes.xlsx')
    ws = wb['arenes']
    ids=[ws.cell(row=i,column=1).value for i in range(2,12)]
    for i, id in enumerate(ids,start=1):
         if id != None :
             nb_arenes = i
    return nb_arenes

#-------------------------------------
#  get tout les arenes 
#------------------------------------    
def get_arenes():
    arenes = []
    

    #recuperer les arenes existantes 
    wb = openpyxl.load_workbook('arenes.xlsx')
    ws = wb['arenes']

    ids=[ws.cell(row=i,column=1).value for i in range(2,12)]
    names=[ws.cell(row=i,column=2).value for i in range(2,12)]
    villes=[ws.cell(row=i,column=3).value for i in range(2,12)]
    dresseurs=[ws.cell(row=i,column=4).value for i in range(2,12)]
    types=[ws.cell(row=i,column=5).value for i in range(2,12)]
    scores=[ws.cell(row=i,column=6).value for i in range(2,12)]
    colores=[ws.cell(row=i,column=7).value for i in range(2,12)]

         
    

    for i, ville in enumerate(villes):
        if ville != None :
           arenes.append(Arene(ids[i],names[i],villes[i],dresseurs[i],types[i],scores[i],colores[i]))        
           

    return arenes
#-------------------------------------
#   afficher les arenes
#------------------------------------
def affiche_arenes(arenes):
    print(termcolor.colored("Dans le monde des Pokemons il existe", "yellow"),len(arenes),termcolor.colored("Arénes : \n", "yellow"))
    termcolors =["red","cyan","green","yellow","grey","magenta","cyan"]
    table = []
    for arene in arenes:
        table.append([
            termcolor.colored(arene.name, arene.color),
            termcolor.colored(arene.ville, arene.color),
            termcolor.colored(arene.dresseur, arene.color),
            termcolor.colored(arene.type, arene.color),
            termcolor.colored(arene.score, arene.color)
        ])
 
    headers = ["Nom", "Ville", "Dresseur","type","score"]

    print(tabulate(table, headers=headers, tablefmt="fancy_grid"))
    
    # for i, arene in enumerate(arenes):
    #     print(termcolor.colored("nom :",arene.color),
    #           termcolor.colored(arene.name,arene.color),
    #           termcolor.colored("|",arene.color),
    #           termcolor.colored("ville",arene.color),
    #           termcolor.colored(arene.ville,arene.color),
    #           termcolor.colored("|",arene.color),
    #           termcolor.colored("dresseur",arene.color),
    #           )
        #print(termcolor.colored(arene.name, termcolors[i % len(termcolors)]))
    print('\n')

#-------------------------------------
#   enregister  un arene
#------------------------------------
def add_arene(arene):    

    wb = openpyxl.load_workbook('arenes.xlsx')
    ws = wb['arenes']
    new_arene = (arene.id,arene.name,arene.ville,arene.dresseur,arene.type,arene.score,arene.color)
    ws.append(new_arene)
    wb.save('arenes.xlsx')
    print(termcolor.colored("✅ Votre Arene est ajouter avec sccces !!!! ","green"))

#-------------------------------------
#   verifier si l'arene existe 
#------------------------------------
def arene_existe(nom):

    arenes = get_arenes()
    list = []
    for i in arenes :
        list.append(i.name)
    if nom in list:
        return True
    else :
        return False


#-------------------------------------
#   cration de l'objet arene
#------------------------------------
def create_bject_arene():
    termcolors =["red","cyan","green","yellow","grey","magenta","cyan"]
    arenes = get_arenes()
    list = []
    for i in arenes :
        list.append(i.name)
    while True:
        nom = input("🪧 entrez le nom : ")
        if nom == "":
            continue
        if arene_existe(nom):
            print(" 🚫 l arene existe déja")
            continue
        else:
            break


    while True:
         ville = input("🏙️  entrez la ville  :")
         if ville == "" :
             continue
         else :
             break

    while True:
         dresseur = input("🧑‍🎓 entrez le dressur  :")
         if dresseur == "":
             continue
         else :
             break

    while True:
         type = input("🔥 |💧 |🌿 |⚡ |❄️ |🪨 | 🦋  entrez le type :") 
         if type == "" : 
             continue
         else :
             break
     
  
    color = random.choice(termcolors)
    print(color)
    arene = Arene(get_nb_arenes()+1,nom, ville, dresseur, type, 0,color)
    return arene

#-------------------------------------
#   get arene by name  
#------------------------------------
def get_arene_by_name(name):
    arenes = get_arenes()
    for a in arenes :
        if a.name == name:
            return a

#-------------------------------------
#   get arenes hors choix
#------------------------------------
def get_reste_arenes(arene_id):
    arenes = get_arenes()
    list_arenes = []

    for a in arenes :
        if a.id != arene_id:
            list_arenes.append(a)
    return list_arenes 






   

    


      
    