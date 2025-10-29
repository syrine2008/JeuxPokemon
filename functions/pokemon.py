import termcolor
import openpyxl
import random
from tabulate import tabulate
from models.Pokemon import Pokemon

def get_nb_pokemons():
    nb_pokemons = 0
    wb = openpyxl.load_workbook('arenes.xlsx')
    ws = wb['pokemons']
    ids = [ws.cell(row=i, column=1).value for i in range(2, 102)]  # Ajuster selon le nb max
    for i, id in enumerate(ids, start=1):
        if id is not None:
            nb_pokemons = i
    return nb_pokemons


def get_pokemons_by_arena_id(arena_id):
    pokemons = []
    wb = openpyxl.load_workbook('arenes.xlsx')
    ws = wb['pokemons']

    ids = [ws.cell(row=i, column=1).value for i in range(2, 102)]
    names = [ws.cell(row=i, column=2).value for i in range(2, 102)]
    arenes = [ws.cell(row=i, column=3).value for i in range(2, 102)]
    scores = [ws.cell(row=i, column=4).value for i in range(2, 102)]
    pvs = [ws.cell(row=i, column=5).value for i in range(2, 102)]   
    attaques = [ws.cell(row=i, column=6).value for i in range(2, 102)]
    niveaux = [ws.cell(row=i, column=7).value for i in range(2, 102)]
    points_bonus = [ws.cell(row=i, column=8).value for i in range(2, 102)]
    double_vies = [ws.cell(row=i, column=19).value for i in range(2, 102)]


    for i, arene in enumerate(arenes):
        if arene is not None and arene == arena_id :
            pokemons.append(
                Pokemon(   
                id=ids[i],    
                name=names[i],
                arene=arenes[i],
                score=scores[i],
                pv=pvs[i],
                attaque=attaques[i],
                niveau=niveaux[i],
                points_bonus=points_bonus[i] ,
                double_vie=bool(double_vies[i])
              )
            )
    return pokemons


def affiche_pokemons(pokemons):
    print(termcolor.colored("Dans cette arène, il existe ", "yellow"),
          len(pokemons),
          termcolor.colored("Pokémons :\n", "yellow"))

    termcolors = ["red", "cyan", "green", "yellow", "grey", "magenta", "cyan"]
    table = []

    for i, p in enumerate(pokemons):
        color = termcolors[i % len(termcolors)]
        table.append([
            termcolor.colored(p.name, color),
            termcolor.colored(p.arene, color),
            termcolor.colored(p.score, color),
            termcolor.colored(p.pv, color),
            termcolor.colored(p.attaque, color),
            termcolor.colored(p.niveau, color),
            termcolor.colored(p.points_bonus, color),
            termcolor.colored(p.double_vie, color)
        ])

    headers = ["Nom", "Arène", "Score", "PV", "Attaque", "Niveau", "Bonus", "Double Vie"]
    print(tabulate(table, headers=headers, tablefmt="fancy_grid"))
    print("\n")


def add_pokemon(pokemon):
    wb = openpyxl.load_workbook('arenes.xlsx')
    ws = wb['pokemons']

    new_pokemon = (
        get_nb_pokemons() + 1,
        pokemon.name,
        pokemon.arene,
        pokemon.score,
        pokemon.pv,
        pokemon.attaque,
        pokemon.niveau,
        pokemon.points_bonus,
        pokemon.double_vie
  
    )

    ws.append(new_pokemon)
    wb.save('arenes.xlsx')
    print(termcolor.colored("✅ Votre Pokémon a été ajouté avec succès !", "green"))


def pokemon_existe(name):
    pokemons = get_pokemons_by_arena_id()
    return any(p.name == name for p in pokemons)


def create_object_pokemon(arene_id):


    while True:
        name = input("🪧 Entrez le nom du Pokémon : ")
        if name == "":
            continue
        else:
            break

    while True:
        pv = int(input("❤️ Entrez les PV : "))
        if pv == "":
            continue
        else : 
            break
    while True:
        attaque = int(input("⚔️ Entrez l'attaque : "))
        if attaque == "":
            continue
        else:
            break
    while True:
        double_vie = input("💖 Double vie ? (oui/non) : ")
        if double_vie != "" or double_vie == "oui" or double_vie == "non" :
            break
        else : 
            continue

    pokemon = Pokemon(get_nb_pokemons()+1,name, arene_id, 0, pv, attaque, 1, 0, double_vie)
    return pokemon

def get_pokemon_by_name(name):
    for p in get_pokemons_by_arena_id():
        if p.name == name:
            return p
    return None
