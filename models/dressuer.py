import openpyxl
import functions.pokemon as pokemonfn

def soigner_pokemon(pokemon):
    wb = openpyxl.load_workbook('arenes.xlsx')
    ws = wb['pokemons']

    for i in range(2,pokemonfn.get_nb_pokemons()):
        id_cell = ws.cell(row=i,column=1).value
        if id_cell == pokemon.id:
            ws.cell(row=i,column=5).value += 10
            pokemon.niveau += 10
            break
    wb.save('arenes.xlsx')
    return pokemon